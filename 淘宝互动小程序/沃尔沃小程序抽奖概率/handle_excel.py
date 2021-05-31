# 处理表格
# encoding=utf-8
import openpyxl  # openpyxl是处理Excel表格的模块
import sys
import os
base_path = os.getcwd()+'/沃尔沃小程序抽奖概率'
sys.path.append(base_path)

# 封装Excel表格的操作


class HandExcel:
    def load_excel(self):
        '''
        加载excel
        '''
        open_value = openpyxl.load_workbook(base_path+'/case/data.xlsx')
        return open_value

    def get_sheet_data(self, index=None):
        '''
         加载某一sheet的所有数据
        '''
        sheet_names = self.load_excel().sheetnames  # 获取所有的sheet
        if index == None:
            index = 0
        data = self.load_excel()[sheet_names[index]]  # 获取某一sheet
        return data

    def get_rows(self):
        '''
        获取行数
        '''
        row = self.get_sheet_data().max_row
        return row

    def get_rows_data(self, row):
        '''
        获取某一行内容
        '''
        row_list = []
        # 获取行对象：self.get_sheet_data()[row]
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)  # 获取行内容：self.get_sheet_data()[row].value
        return row_list

    def excel_write_data(self, rows, cols, value):
        '''写入数据，即执行结果回写'''
        wb = self.load_excel()  # 获取表格
        wr = wb.active  # 激活表格，然后可写入
        wr.cell(rows, cols, value)  # 写入数据
        wb.save(base_path+"/case/data.xlsx")  # 保存数据

    def get_excel_data(self):
        '''
        获取excel所有数据     #方便后面数据驱动拿数据
        '''
        data_list = []

        for i in range(1, self.get_rows()):
            data_list.append(self.get_rows_data(i+1))
        return data_list


# 实例化类
excel_data = HandExcel()
